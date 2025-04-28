from openpyxl import load_workbook
import smtplib
from email.mime.text import MIMEText
from email.header import Header
#加载excel
wb=load_workbook('xm22gongzi.xlsx')
#获取当前激活的工作薄
sheet1=wb.active
#获取行数据
count=0
#记录表头
th=''
for row in sheet1.iter_rows():
    if count==0:
        th= f"""
        <tr>
            <td>{row[0].value}</td>
            <td>{row[1].value}</td>
            <td>{row[2].value}</td>
            <td>{row[3].value}</td>
            <td>{row[4].value}</td>
            <td>{row[5].value}</td>
            <td>{row[6].value}</td>
            <td>{row[7].value}</td>
            <td>{row[8].value}</td>
            <td>{row[10].value}</td>
        </tr>
        """
        count+=1
        continue

        # msg=f'id:{row[0].value} name:{row[1].value} name:{row[2].value} 部门:{row[3].value} 基本工资:{row[4].value} 应发工资:{row[9].value}'
    msg= f"""
        <tr>
            <td>{row[0].value}</td>
            <td>{row[1].value}</td>
            <td>{row[2].value}</td>
            <td>{row[3].value}</td>
            <td>{row[4].value}</td>
            <td>{row[5].value}</td>
            <td>{row[6].value}</td>
            <td>{row[7].value}</td>
            <td>{row[9].value}</td>
            <td>{row[10].value}</td>
        </tr>
        """
    count+=1
    to_addr=row[10].value
    #发送协议
    #邮件的地址服务器是哪个
    smtp_obj=smtplib.SMTP_SSL('smtp.qq.com',465)
    #登录邮箱
    smtp_obj.login('3324097560@qq.com','nfnnhghhrtsbchcd')
    #定义发送内容
    msg_body=MIMEText('<table border="1">'+th+msg+'</table>','html','utf-8')
    #定义从哪发来的邮件
    msg_body['From']='Hhhh <3324097560@qq.com>'
    #邮件主题
    msg_body['Subject']=Header('第一次发送邮件','utf-8')
    #发送邮件
    smtp_obj.sendmail('3324097560@qq.com','taulyl_sc@outlook.com',msg_body.as_string())
    print(f'已成功发送邮件给 名为：{row[1].value} 邮箱地址：{to_addr}')

